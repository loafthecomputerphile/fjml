import xlwings as xw
import openpyxl 
from functools import lru_cache

from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from typing import (
    Optional,
    Union, 
    Any,
    Type,
    NoReturn,
    TypeAlias
)

import asyncio

from psutil import (
    process_iter, 
    Process, 
    NoSuchProcess
)
import re
import numpy as np
import pandas as pd
from dataclasses import dataclass, field
from abc import ABCMeta, abstractmethod


class VariableFileError(Exception):
    pass


@dataclass
class RefreshedCell:
    row: int = field(default=0)
    col: int = field(default=0)
    value: str = field(default="") 


class BookUtils:
    
    @lru_cache(None)
    def compare(self, old: str, new: str) -> bool:
        return new == old
    
    def get_changes(self, old_data: pd.DataFrame, new_data: pd.DataFrame, columns: list[str]) -> list[RefreshedCell]:
        idx: int
        row: pd.Series
        cell: str
        col_name: str
        temp_col: list[str]
        result: list[RefreshedCell] = []
        
        df: pd.DataFrame = old_data.compare(new_data)
        df.drop("self", axis=1, level=1, inplace=True)
        df.columns = df.columns.droplevel(1)
        
        for row_idx, row in df.iterrows():
            temp_col = columns
            for idx, (cell, col_name) in enumerate(zip(row, row.index)):
                #this allows the distinguishment of duplicate column values
                if col_name in columns[:idx]:
                    temp_col[:idx] = [-1 for _ in range(idx)]
                    
                result.append(
                    RefreshedCell(
                        row=row_idx, 
                        col=temp_col.index(col_name), 
                        value=cell
                    )
                )
        
        return result

    def is_formula(self, data: str) -> bool:
        return (
            data[0] == "="
        )
    
    def letters_to_index(self, letters: str) -> int:
        uppercase_letters: str = letters.upper()
        if len(uppercase_letters) == 1:
            return (ord(uppercase_letters) - 64)
        
        index: int = 0
        for letter in uppercase_letters:
            index += (ord(letter) - 64)
        return index
    
    def difference(self, list1: list[str], list2: list[str]) -> list[str]:
        return list(
            set(list1).difference(list2)
        )


@dataclass
class Filter:
    not_in: list[str] = field(default=None)
    is_in: list[str] = field(default=None)
    
    def __post_init__(self) -> NoReturn:
        if self.not_in and self.is_in:
            raise Error("One filter must be used at a time")
    
    def apply_filter(self, sheets: str) -> list[str]:
        sheet: str
        
        return [
            sheet for sheet in sheets if self.__any(sheet)
        ]
    
    def __any(self, sheet: str) -> bool:
        item: str
        
        if self.not_in:
            return any(item not in sheet for item in self.not_in)
        
        if self.is_in:
            return any(item in sheet for item in self.is_in)


@dataclass
class Sheet:
    title_row: int
    start_row: int
    end_row: int
    column_letters: list[str] = field(default_factory=list)
    editable_col_letters: list[str] = field(default_factory=list)
    column_dtypes: list[dict] = field(default_factory=list)


@dataclass
class Table:
    table_names: list[str] = field(default_factory=list)
    column_letters: list[list[str]] = field(default_factory=list)
    editable_col_letters: list[list[str]] = field(default_factory=list)
    column_dtypes: list[dict] = field(default_factory=list)
    title_rows: list[int] = field(default_factory=list)
    start_rows: list[int] = field(default_factory=list)
    end_rows: list[int] = field(default_factory=list)
    
    def __post_init__(self) -> NoReturn:
        names_length: int = len(self.table_names)
        start_rows_length: int = len(self.start_rows)
        letters_length: int = len(self.column_letters)
        end_rows_length: int = len(self.end_rows)
        title_row_length: int = len(self.title_rows)
        edit_col_length: int = len(self.editable_col_letters)
        if not (
            names_length == start_rows_length and start_rows_length == letters_length and
            letters_length == end_rows_length and end_rows_length == title_row_length and
            title_row_length == edit_col_length
        ):
            raise VariableFileError("length of every setting should be the same")


@dataclass
class Value:
    itype: str
    model: Union[Sheet, Table]


@dataclass
class TableParams:
    sheets: Union[list[str], str] = field(default=None)
    sheet_filter: dict[str, list[str]] = field(default=None)
    settings: dict[str, dict] = field(default_factory=dict)
    all_sheets: list[str] = field(default_factory=list)
    
    def __post_init__(self) -> NoReturn:
        self.params: dict[str, Value] = dict()
        if self.sheets:
            if self.sheet_filter:
                raise Error("sheet_filter can not be assigned when sheets is assigned also")
            if isinstance(self.sheets, str):
                if self.sheet != "remaining":
                    raise Error(f"""
                        if sheets value is a String the only 
                        acceptable value is 'remaining', recieved '{self.sheets}'
                    """)
                self.params["remaining"] = Value("table", Table(**self.settings))
                self.sheets = self.all_sheets
            elif isinstance(self.sheets, list):
                self.params = {
                    sheet: Value("table", Table(**self.settings)) for sheet in self.sheets if sheet in self.all_sheets
                }
            return
        
        if self.sheet_filter:
            self.filter: Filter = Filter(**self.sheet_filter)
            self.sheets = self.run_filter()
            self.params = {
                sheet: Value("table", Table(**self.settings)) for sheet in self.sheets if sheet in self.all_sheets
            }
        
    def run_filter(self) -> list[str]:
        return self.filter.apply_filter(
            self.all_sheets
        )


@dataclass
class SheetParams(TableParams):
    
    def __post_init__(self) -> NoReturn:
        self.filter: Filter
        self.params: dict[str, Value] = dict()
        if self.sheets:
            if self.sheet_filter:
                raise Error("sheet_filter can not be assigned when sheets is assigned also")
            if isinstance(self.sheets, str):
                if self.sheets != "remaining":
                    raise Error(f"""
                        if sheets value is a String the only 
                        acceptable value is 'remaining', recieved '{self.sheets}'
                    """)
                self.params["remaining"] = Value("sheet", Sheet(**self.settings))
                self.sheets = self.all_sheets
                return
            elif isinstance(self.sheets, list):
                self.params = {
                    sheet: Value("sheet", Sheet(**self.settings)) for sheet in self.sheets if sheet in self.all_sheets
                }
            return
        
        if self.sheet_filter:
            self.filter = Filter(**self.sheet_filter)
            self.sheets = self.run_filter()
            self.params = {
                sheet: Value("sheet", Sheet(**self.settings)) for sheet in self.sheets if sheet in self.all_sheets
            }


ParamType = Union[Type[SheetParams], Type[TableParams]]


@dataclass
class JoinAll:
    sheet_data: Union[list[dict[str, Any]], None] = field(default=None)
    table_data: Union[list[dict[str, Any]], None] = field(default=None)
    all_sheets: list[str] = field(default=list)
    
    def __post_init__(self) -> NoReturn:
        self.params: dict[str, Value] = dict()
        self.used_sheets = set()
        success: bool
        
        if self.sheet_data:
            success = self.generate_paramaters(SheetParams, self.sheet_data)
            if not success:
                return
        
        if self.table_data:
            success = self.generate_paramaters(TableParams, self.table_data)
            if not success:
                return
    
    def difference(self, set1: set[str], set2: set[str]) -> list[str]:
        return list(
            set1.difference(set2)
        )
        
    def generate_paramaters(self, param_class: ParamType, spread_type: list[dict[str, Any]]) -> bool:
        data: dict[str, Any]
        param_model: Union[SheetParams, TableParams]
        model: Value
        sheet: str
        
        for data in spread_type:
            param_model = param_class(
                data["sheets"], data["sheet_filter"], 
                data["settings"], self.all_sheets
            )
            if data["sheets"] == "remaining":
                model = param_model.params["remaining"]
                for sheet in self.difference(set(self.all_sheets), self.used_sheets):
                    self.params[sheet] = model
                return False
            else:
                self.params.update(param_model.params)
                self.used_sheets.update(data["sheets"])
        return True


JsonInput = Optional[list[dict[str, Any]]]


class BaseBook(metaclass=ABCMeta):
    
    __slots__: tuple[str, ...] = (
        "start_row", "end_row", "restricted_sheets",
        "title_column", "column_letters", "editable_col_letters",
        "editable_col_length", "current_row", "current_sheet_name",
        "sheet", "book", "app",
        "is_loaded", "table_indexes", "file_path",
        "last_input_row", "tools"
    )
    
    def __init__(
        self, table_params: JsonInput = None, sheet_params: JsonInput = None, 
        restricted_sheets: list[str] = []
    ) -> NoReturn:
        self.table_params: JsonInput = table_params
        self.sheet_params: JsonInput = sheet_params
        self.tools: BookUtils = BookUtils()
        self.main_data: dict[str, Value]
        self.start_row: int
        self.end_row: int
        self.restricted_sheets: list[str]
        self.title_row: int
        self.column_letters: list[str]
        self.editable_col_letters: list[str]
        self.editable_col_length: int
        self.current_row: int
        self.current_sheet_name: str
        self.file_path: str
        self.sheet: Union[xw.Sheet, worksheet.Worksheet]
        self.book: Union[xw.Book, openpyxl.Workbook]
        self.table: pd.DataFrame
        self.app: xw.App
        self.is_loaded: bool
        self.table_names: list[str]
        self.table_idx: int
        self.in_table: bool
        #self.table_inclusive_sheets = self.sheet_tables.is_sheets_inclusive
    
    #@abstractmethod
    async def select_table(self, table_name: str = None) -> NoReturn:
        raise NotImplementedError()
    
    async def load_book(
        self, file_path: str, 
        restricted_sheets: Optional[list[str]] = None
    ) -> NoReturn:
        self.restricted_sheets = restricted_sheets
        self.main_data: dict[str, Value] = JoinAll(
            self.sheet_params,
            self.table_params,
            self.tools.difference(
                self.sheet_names,
                self.restricted_sheets
            )
        ).params
    
    async def select_table(self, table_name: str = None) -> pd.DataFrame:
        if table_name:
            
            data: Union[Value, None] = self.main_data.get(
                self.current_sheet_name, 
                None
            )
            model: Union[Sheet, Table] = data.model
            
            self.table_idx = self.table_names.index(table_name)
            self.start_row = model.start_rows[self.table_idx]
            self.end_row = model.end_rows[self.table_idx]
            self.title_row = model.title_rows[self.table_idx]
            self.column_letters = model.column_letters[self.table_idx]
            self.editable_col_letters = model.editable_col_letters[self.table_idx]
            self.editable_col_length = len(self.editable_col_letters)
            self.current_row = self.start_row
            self.current_table_name = table_name
            
            self.table = await self.get_sheet_data()
        
        return self.table
    
    async def select_sheet(self, sheet_name: str) -> NoReturn:
        if sheet_name in self.restricted_sheets:
            return 
        
        data: Union[Value, None] = self.main_data.get(sheet_name, None)
        
        if data == None:
            return
        
        model: Union[Sheet, Table] = data.model
        input_type: str = data.itype
        
        if input_type == "sheet":
            self.table_idx = None
            self.start_row = model.start_row
            self.end_row = model.end_row
            self.title_row = model.title_row
            self.column_letters = model.column_letters
            self.editable_col_letters = model.editable_col_letters
            self.editable_col_length = len(self.editable_col_letters)
            self.current_row = self.start_row
            self.table_names = []
            self.in_table = False
        elif input_type == "table":
            self.table_idx = 0
            self.start_row = model.start_rows[self.table_idx]
            self.end_row = model.end_rows[self.table_idx]
            self.title_row = model.title_rows[self.table_idx]
            self.column_letters = model.column_letters[self.table_idx]
            self.editable_col_letters = model.editable_col_letters[self.table_idx]
            self.editable_col_length = len(self.editable_col_letters)
            self.current_row = self.start_row
            self.table_names = model.table_names
            self.in_table = True
        
        self.current_sheet_name = sheet_name
    
    
    async def refresh(self, data: pd.DataFrame) -> list[RefreshedCell]:
        return self.tools.get_changes(
            old_data=data, 
            new_data=(
                await self.get_sheet_data()
            ), 
            columns=self.column_names
        )
    
    #@abstractmethod
    async def add_rows(self, num: int) -> NoReturn:
        raise NotImplementedError()
    
    #@abstractmethod
    async def append_row(self, data: list[Any], refresh: bool) -> list[Any]:
        raise NotImplementedError()
    
    #@abstractmethod
    async def get_row_data(self, row=None) -> np.ndarray:
        raise NotImplementedError()
    
    #@abstractmethod
    async def get_sheet_data(
        self, end_row: Optional[int] = None, 
        letters: list[str] = None, start_row: Optional[int] = None
    ) -> pd.DataFrame:
        raise NotImplementedError()
    
    #@abstractmethod
    async def switch_sheet(self, sheet_name: str) -> pd.DataFrame:
        raise NotImplementedError()
    
    @property
    #@abstractmethod
    def sheet_names(self) -> list[str]:
        raise NotImplementedError()
    
    @property
    #@abstractmethod
    def last_used_row(self) -> int:
        raise NotImplementedError()
    
    @property
    def column_names(self) -> list[str]:
        return asyncio.run(
            self.get_row_data(
                self.title_row
            )
        )
    
    def increment_currrent_row(self) -> NoReturn:
        if self.current_row < self.end_row:
            self.current_row+=1 
    
    def change_row(self, row) -> NoReturn:
        if row <= self.end_row and row >= self.start_row:
            self.current_row = row
    
    async def close_excel_files(self) -> NoReturn:
        p: Optional[Process] = None
        for process in process_iter(['pid', 'name']):
            if process.info['name'] == 'EXCEL.EXE': # type: ignore
                try:
                    p = Process(process.info['pid']) # type: ignore
                    p.terminate()    # Terminate the Excel process
                    p.wait()             # Wait for the process to terminate
                except e:
                    pass 

    #@abstractmethod
    async def close(self) -> NoReturn:
        raise NotImplementedError()
        
    #@abstractmethod
    async def save(self) -> NoReturn:
        raise NotImplementedError()


class OpenpyxlBookModel(BaseBook):

    async def load_book(
        self, file_path: str, 
        restricted_sheets: Optional[list[str]] = []
    ) -> NoReturn:
        self.file_path = file_path
        self.book = openpyxl.load_workbook(file_path)
        await super().load_book(file_path, restricted_sheets)
        self.is_loaded = True
    
    async def select_sheet(self, sheet_name: str) -> NoReturn:
        await super().select_sheet(sheet_name)
        self.sheet = self.book.get_sheet_by_name(sheet_name)

    async def add_rows(self, num: int) -> NoReturn:
        self.sheet.insert_rows(self.current_row, amount=num)
    
    async def insert_data(self, data: pd.DataFrame, row_idx: Optional[int] = None) -> list[RefreshedCell]:
        row_idx: int = row_idx if row_idx else self.current_row
        
        for col in data.columns:
            letter: str = self.column_letters[self.column_names.index(col)]
            for idx, value in enumerate(data[col]):
                self.sheet[f"{letter}{(row_idx+idx)}"] = value
        
        return await self.refresh(df)
    
    async def append_row(self, data: list[str], df: pd.DataFrame) -> list[RefreshedCell]:
        i: int
        value: str
        cell: str
        
        for i, value in enumerate(data):
            cell = f"{self.editable_col_letters[i]}{self.current_row}"
            self.sheet[cell] = value
        
        return await self.refresh(df)

    async def get_row_data(self, row: Optional[int] = None) -> list[str]:
        col: str
        row = row if row else self.current_row
        return [
            self.sheet[f"{col}{row}"].value 
            for col in self.column_letters
        ]

    async def get_sheet_data(self) -> pd.DataFrame:
        letter_indexes: list[int] = [self.tools.letters_to_index(i) for i in self.column_letters]
        max_idx: int = max(letter_indexes)
        i: int
        
        df: pd.DataFrame = pd.DataFrame(
            self.sheet.iter_rows(
                values_only=True, min_row=self.start_row, 
                max_col=max_idx, max_row=self.end_row,
            ),
            columns=[i+1 for i in range(max_idx)]
        )[letter_indexes]
        '''
        new_col = self.column_names
        m = {f"{column}": f"{new_col[i]}" for i, column in enumerate(df.columns)}
        '''
        df.columns = self.column_names
        #print(df.rename(columns=m))
        return df

    async def switch_sheet(self, sheet_name: str) -> pd.DataFrame:
        await self.select_sheet(sheet_name)
        if self.in_table:
            return await self.select_table()
        return await self.get_sheet_data()
    
    @property
    def sheet_names(self) -> list[str]:
        return self.book.sheetnames

    @property
    def last_used_row(self) -> int:
        return self.sheet.max_row

    async def close(self) -> NoReturn:
        if self.book:
            self.book.close()
        self.is_loaded = False

    async def save(self, path: Optional[str] = None) -> NoReturn:
        if self.book:
            self.book.save(path)


class XlwingsBookModel(BaseBook):
    
    async def load_book(
        self, file_path: str, 
        restricted_sheets: Optional[list[str]] = []
    ) -> NoReturn:
        #await self.close_excel_files()
        self.app = xw.App(visible=False)
        self.book = xw.Book(file_path, mode="i")
        self.app.screen_updating = False
        self.visible = False
        
        await super().load_book(file_path, restricted_sheets)
        self.is_loaded = True
    
    async def select_sheet(self, sheet_name: str) -> NoReturn:
        await super().select_sheet(sheet_name)
        self.sheet = self.book.sheets[sheet_name]
    
    async def insert_data(self, data: pd.DataFrame, row_idx: Optional[int] = None) -> list[RefreshedCell]:
        row_idx: int = row_idx if row_idx else self.current_row
        row_count: int = data.shape[1]
        
        for col, letter in zip(self.column_names, self.column_letters):
            self.sheet[f"{letter}{row_idx}"] = data[col].to_numpy().reshape(-1, 1)

        return await self.refresh(df)
    
    def add_rows(self, selected_end_row: int, num: int) -> NoReturn:
        last_row = self.current_row+selected_end_row
        for col_letter in self.editable_col_letters:
            col_range: str = f'{col_letter}{self.current_row}:{col_letter}{last_row}'
            
            # Insert rows for the entire column
            self.sheet.range(col_range).api.EntireColumn.Insert(
                Shift=xw.constants.InsertShiftDirection.xlShiftDown
            )
    
    async def append_row(self, data: list[Any], df: pd.DataFrame) -> list[Any]:
        if len(data) == self.editable_col_length:
            for col, val in zip(self.editable_col_letters, data):
                if BookUtils.is_formula(val):
                    self.sheet.range(f'{col}{self.current_row}').formula = val
                else:
                    self.sheet.range(f'{col}{self.current_row}').value = val
        else:
            raise Exception(
                "length of array 'data' must be the same length as 'self.editable_col_letters'. "
            )
            
        return await self.refresh(df)

    async def get_row_data(self, row: Optional[int] = None) -> list[str]:
        row = row if row else self.current_row
        return [
            self.sheet.range(f'{col}{row}').value 
            for col in self.column_letters
        ]
    
    async def get_sheet_data(self) -> pd.DataFrame:
        letters: list[str] = self.column_letters
        letter_indexes: list[int] = [self.tools.letters_to_index(i)-1 for i in letters]
        
        df: pd.DataFrame = self.sheet.range(
            (self.start_row, min(letters)),
            (self.end_row, max(letters))
        ).options(
            pd.DataFrame, 
            chunksize=10_000
        ).value[letter_indexes]
        
        df.columns = self.column_names
        '''
        df = df.rename(columns={
            f"{column}": f"{new_col[i]}" for i, column in enumerate(df.columns)
        })
        '''
        return df
    
    async def switch_sheet(self, sheet_name: str) -> pd.DataFrame:
        await self.select_sheet(sheet_name=sheet_name)
        if self.in_table:
            return await self.select_table()
        return await self.get_sheet_data()
    
    @property
    def sheet_names(self) -> list[str]:
        return self.book.sheet_names
    
    @property
    def last_used_row(self) -> int:
        used_address: str = self.sheet.used_range.address
        return int(used_address.split("$")[-1])

    #closes the book and applications opened by the program
    async def close(self) -> NoReturn:
        if self.book:
            self.book.close()
        
        if self.app:
            self.app.screen_updating = True
            self.app.quit()# type: ignore
        
        self.is_loaded = False

    #saves the book to the test folder or to the edited file
    async def save(self) -> NoReturn:
        self.book.save(None)

    @property
    def visible(self) -> bool:
        return self.app.visible
    
    @visible.setter
    def visible(self, value: bool) -> NoReturn:
        self.app.visible = value


class BookFactory:
    def create_book(self, class_kwargs: dict) -> Union[OpenpyxlBookModel, XlwingsBookModel]:
        
        book_types: dict[str, Union[Type[OpenpyxlBookModel], Type[XlwingsBookModel]]] = {
            "xlwings":XlwingsBookModel,
            "openpyxl":OpenpyxlBookModel
        }
        engine: str = class_kwargs["engine"]
        book_model: Union[OpenpyxlBookModel, XlwingsBookModel] = book_types.get(
            engine, None
        )
        del class_kwargs["engine"]
        
        if book_model:
            return book_model(**class_kwargs)
        else:
            raise ValueError(f"Invalid value for 'engine'. recieved {engine}, expected 'xlwings' or 'openpyxl'")

